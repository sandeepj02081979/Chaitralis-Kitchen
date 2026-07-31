#!/usr/bin/env python3
"""Convert data/menu.xlsx -> data/menu.json for the Chaitrali's Kitchen website.

Run from the repo root:  python3 scripts/build_menu.py
Requires: openpyxl  (pip install openpyxl)
"""
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "data" / "menu.xlsx"
OUT = ROOT / "data" / "menu.json"


def truthy(v):
    return str(v).strip().upper() == "TRUE"


def rows_after_header(ws, header_row, n_cols):
    """Yield rows below the header row, stopping at the first blank row (notes live below the table)."""
    for row in ws.iter_rows(min_row=header_row + 1, max_col=n_cols, values_only=True):
        if row[0] is None or str(row[0]).strip() in ("", "—"):
            break
        yield [("" if c is None else str(c).strip() if not isinstance(c, (int, float)) else c) for c in row]


def build():
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    # Tiffin Menu: header on row 6, week-of in B4
    ws = wb["Tiffin Menu"]
    tiffin = {
        "week_of": str(ws["B4"].value or "").strip(),
        "days": [
            {
                "day": r[0], "bhaji": r[1], "amti": r[2], "rice": r[3],
                "poli": r[4], "side": r[5], "sweet": r[6], "notes": r[7],
            }
            for r in rows_after_header(ws, 6, 8)
        ],
    }

    # Catering Menu: header on row 4
    ws = wb["Catering Menu"]
    catering = [
        {
            "name": r[0], "category": r[1], "price": r[2], "unit": r[3],
            "description": r[4], "spicy": truthy(r[5]), "contains": r[6],
        }
        for r in rows_after_header(ws, 4, 9)
        if truthy(r[7])  # Available
    ]

    # Seasonal Specials: header on row 4
    ws = wb["Seasonal Specials"]
    seasonal = [
        {
            "occasion": r[0], "name": r[1], "price": r[2], "unit": r[3],
            "description": r[4], "order_by": r[5],
        }
        for r in rows_after_header(ws, 4, 8)
        if truthy(r[6])  # Active
    ]

    # Reviews: header on row 4
    ws = wb["Reviews"]
    reviews = [
        {"name": r[0], "location": r[1], "rating": int(r[2]), "text": r[3], "date": r[4]}
        for r in rows_after_header(ws, 4, 5)
    ]

    # Settings: header on row 4, key/value
    ws = wb["Settings"]
    settings = {r[0]: str(r[1]) for r in rows_after_header(ws, 4, 3)}

    data = {
        "settings": settings,
        "tiffin": tiffin,
        "catering": catering,
        "seasonal": seasonal,
        "reviews": reviews,
    }
    OUT.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Wrote {OUT.relative_to(ROOT)}: "
          f"{len(tiffin['days'])} tiffin days, {len(catering)} catering items, "
          f"{len(seasonal)} active seasonal, {len(reviews)} reviews, {len(settings)} settings")


if __name__ == "__main__":
    build()
